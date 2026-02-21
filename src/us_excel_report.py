import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os

DATA_PATH = "../data/processed/macro_us_with_risk.csv"
SNAP_PATH = "../data/processed/sql_snapshot.csv"
LAST12_PATH = "../data/processed/sql_last12.csv"
OUTPUT_FILE = "../reports/macro_us_report.xlsx"
TOP_HIGH_PATH = "../data/processed/top_high_stress_periods.csv"
TOP_LOW_PATH = "../data/processed/top_low_stress_periods.csv"
INDEX_PATH = "../data/processed/macro_us_with_index.csv"

def style_overview(ws):
    # Highlight risk_level cell(s)
    for row in ws.iter_rows(min_row=2, max_row=2):
        for cell in row:
            if cell.value == "HIGH RISK":
                cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif cell.value == "MEDIUM RISK":
                cell.fill = PatternFill(start_color="FFA500", fill_type="solid")
                cell.font = Font(bold=True)
            elif cell.value == "LOW RISK":
                cell.fill = PatternFill(start_color="00B050", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

def create_excel_report():
    os.makedirs("../reports", exist_ok=True)

    df = pd.read_csv(DATA_PATH)
    df["date"] = pd.to_datetime(df["date"])
    latest = df.sort_values("date").iloc[-1:]
    # Optional new insight tabs
    top_high = pd.read_csv(TOP_HIGH_PATH) if os.path.exists(TOP_HIGH_PATH) else None
    top_low = pd.read_csv(TOP_LOW_PATH) if os.path.exists(TOP_LOW_PATH) else None
    index_df = pd.read_csv(INDEX_PATH) if os.path.exists(INDEX_PATH) else None

    current_regime = None
    if index_df is not None:
        index_df["date"] = pd.to_datetime(index_df["date"])
        current_regime = index_df.sort_values("date").iloc[-1:][
            ["date","macro_stress_index","stress_level","macro_strategy"]
        ]

    # Optional SQL tabs
    snap = pd.read_csv(SNAP_PATH) if os.path.exists(SNAP_PATH) else None
    last12 = pd.read_csv(LAST12_PATH) if os.path.exists(LAST12_PATH) else None

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        latest.to_excel(writer, sheet_name="Overview", index=False)
        df.to_excel(writer, sheet_name="Full_Data", index=False)

        if snap is not None:
            snap.to_excel(writer, sheet_name="SQL_Snapshot", index=False)
        if last12 is not None:
            last12.to_excel(writer, sheet_name="SQL_Last12", index=False)
        if current_regime is not None:
            current_regime.to_excel(writer, sheet_name="Current_Regime", index=False)

        if top_high is not None:
            top_high.to_excel(writer, sheet_name="Top_High_Stress", index=False)

        if top_low is not None:
            top_low.to_excel(writer, sheet_name="Top_Low_Stress", index=False)
        

    wb = load_workbook(OUTPUT_FILE)
    style_overview(wb["Overview"])
    wb.save(OUTPUT_FILE)

    print("✅ Excel Report Generated:", OUTPUT_FILE)

if __name__ == "__main__":
    create_excel_report()
